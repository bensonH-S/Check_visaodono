"""Lê xlsx e exporta JSON para o seed de metas."""
import json
import sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

path = sys.argv[1]

wb = openpyxl.load_workbook(path, data_only=True)
out = {'sheets': {}}

for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    out['sheets'][name] = {'rows': rows}

print(json.dumps(out, ensure_ascii=False, default=str))
